from math import floor
import time
import cv2
import mediapipe as mp
import openpyxl
from monitoring import monitoring

mpFaceMesh = mp.solutions.face_mesh
mpDraw = mp.solutions.drawing_utils
faceMesh = mpFaceMesh.FaceMesh(refine_landmarks = True)
drawing_spec = mpDraw.DrawingSpec(thickness=1, circle_radius=1)


def main():

    facemesh_model = mp.solutions.face_mesh.FaceMesh(
        max_num_faces=1,
        refine_landmarks=True,
        min_detection_confidence=0.5,
        min_tracking_confidence=0.5,
    )

    # facemesh_model = get_mediapipe_app()
    eye_idxs = {
        "left": [362, 385, 387, 263, 373, 380],
        "right": [33, 160, 158, 133, 153, 144],
    }
    thresholds = {
        "EAR_THRESH": 0.25,
        "WAIT_TIME": 1.0
    }

    state_tracker ={
        "start_time": time.perf_counter(),
        "DROWSY_TIME": 0.0,
        "COLOR": (0, 255, 0), 
        "EAR_blink": 0.2,
        "blink_rate": 0,
        "blinkCount": 0,
        "flag": False,
    }

    thresholds = {
        "EAR_THRESH": 0.25,
        "WAIT_TIME": 1.0
    }

    # cap = cv2.VideoCapture(0)
    cap = cv2.VideoCapture('Videos/GX010169.MP4')

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret:
            continue
        frame = cv2.resize(frame, (960, 540))
        frame = cv2.flip(frame, 1)
        frame, output = monitoring(frame, facemesh_model, eye_idxs, state_tracker, thresholds)
        cv2.imshow('Drowsiness Detection', frame)
        
        if cv2.waitKey(1) == ord('q'):
            break
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active


    # Write the headers
    sheet.cell(row=1, column=1).value = "Timestamp"
    sheet.cell(row=1, column=2).value = "EAR"
    sheet.cell(row=1, column=3).value = "Blink Rate"
    sheet.cell(row=1, column=4).value = "Blink Count"
    sheet.cell(row=1, column=5).value = "Drowsy Time(s)"
    sheet.cell(row=1, column=6).value = "Region"
    sheet.cell(row=1, column=7).value = "Fixation Time(s)"

    # Write the captured data to the Excel sheet
    for i, (timestamp, ear, br, bc, dt, r, ft) in enumerate(output, start=2):
        sheet.cell(row=i, column=1).value = timestamp
        sheet.cell(row=i, column=2).value = ear
        sheet.cell(row=i, column=3).value = br
        sheet.cell(row=i, column=4).value = bc
        sheet.cell(row=i, column=5).value = dt
        sheet.cell(row=i, column=6).value = r
        sheet.cell(row=i, column=7).value = ft

    # Save the workbook
    workbook.save('position_data2.xlsx')

    print("Data exported to 'position_data2.xlsx'")

    cap.release()
    cv2.destroyAllWindows()


if __name__ == "__main__":
    main()

